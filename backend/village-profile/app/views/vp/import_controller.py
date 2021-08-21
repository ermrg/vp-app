
from datetime import date, datetime

from django.shortcuts import render, redirect
import openpyxl
from pymysql import NULL

from app.models import AnimalType, HouseType, BusinessType
from app.models.vp.house_detail import HouseDetail
from app.models.vp.household import Household
from app.models.vp.household_animal import HouseholdAnimal
from app.models.vp.household_business import HouseholdBusiness
from app.models.vp.household_helper import HouseholdHelper
from app.models.vp.household_education import HouseholdEducation
from app.models.vp.household_expenses import HouseholdExpenses
from app.models.vp.household_facility import HouseholdFacility
from app.models.vp.household_festival import HouseholdFestival
from app.models.vp.household_fuel import HouseholdFuel
from app.models.vp.household_healthpost import HouseholdHealthPost
from app.models.vp.household_income import HouseholdIncome
from app.models.vp.household_infrastructure import HouseholdInfrastructure
from app.models.vp.household_land import HouseholdLand
from app.models.vp.household_natural_disaster import HouseholdNaturalDisaster
from app.models.vp.household_public_transport import HouseholdPublicTransport
from app.models.vp.household_sewage import HouseholdSewage
from app.models.vp.household_watersource import HouseholdWaterSource
from app.models.vp.member import Member
from app.models.vp.member_deceased import MemberDeceased
from app.models.vp.rent_member import RentMember


def import_excel(request):
    return render(request, 'app/backend/vp/import.html')


def upload_excel(request):
    excel_file = request.FILES["excel_file"]
    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets

    households = wb["Sheet1"]
    expire_members = wb["data-Expire-Member-Details"]
    members = wb["data-House-Member-Details"]
    rent_details = wb["data-rental-details"]
    helper_details = wb["data-household-helper-details"]
    land_details = wb["data-land-details"]
    animal_details = wb["data-Household-animal-info"]

    # sheets = wb.sheetnames
    # print(sheets)  # this prints all sheets heading
    #
    for i, house in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        household = Household()
        household.id_string = house[0].value
        household.ward_id = house[1].value
        household.basti_id = house[2].value
        # household.marga = house[3].value
        household.religion_id = house[4].value
        household.jaati_id = house[5].value
        household.mother_tongue_id = house[6].value
        household.main_occupation_id = house[7].value
        household.resident_type = house[8].value
        # household.ownership = house[8].value
        household.year_if_migrated = house[9].value
        household.phone_num = house[10].value
        household.house_num = house[11].value
        household.num_of_member = house[12].value
        household.hoh_name = house[13].value

        # geo_location = house[14].value  # commented since odk form was getting the location
        # household_location = geo_location.split(",")
        # household.latitude = household_location[0]
        # household.longitude = household_location[1]

        # household.altitude = house[15].value
        # household.accuracy = house[16].value

        household.responder_name = house[17].value
        household.responder_image = house[18].value

        household.status = 1
        household.save()

        # print("House Detail From Sheet1")
        # print(house[0].value)  # data-meta-instanceID
        # print(house[1].value)  # data-Household-info-Household-Info-1-Ward-No
        # print(house[2].value)  # data-Household-info-Household-Info-1-Basti-Name
        # print(house[3].value)  # data-Household-info-Household-Info-1-Other-Basti-Name
        # print(house[4].value)  # data-Household-info-Household-Info-1-Marga-No
        # print("===================")
        # #  Save houshold here

    # for i, deceased_member in enumerate(expire_members.iter_rows(min_row=2)):  # For member sheet
    #     member_dead = MemberDeceased()
    #     member_dead.id_string = deceased_member[8].value
    #     member_dead.name = deceased_member[1].value
    #     hh_id_string = deceased_member[7].value
    #     hh = Household.objects.filter(id_string=hh_id_string).first()
    #     member_dead.hh = hh
    #     member_dead.hh_id_string = hh_id_string
    #     member_dead.gender_id = deceased_member[2].value
    #     #  member_dead.ward_id = house[1].value
    #     member_dead.reason_of_death_id = deceased_member[5].value
    #     member_dead.age_on_death = deceased_member[3].value
    #     member_dead.month_of_death = deceased_member[4].value
    #     member_dead.save()

    for i, member in enumerate(members.iter_rows(min_row=2)):  # For member sheet

        member_details = Member()
        member_details.id_string = member[50].value
        hh_id_string = member[49].value
        hh = Household.objects.filter(id_string=hh_id_string).first()
        member_details.hh = hh
        member_details.name_nep = member[0].value
        member_details.relation_with_hoh_id = member[1].value
        if member[1].value == 1:
            member_details.is_hoh = 1
        else:
            member_details.is_hoh = 0
        # date_str = datetime(member[4].value)
        date_str = str(member[4].value)
        recorded_date = datetime.strptime(date_str, '%d/%m/%y').date()
        today = date.today()
        member_details.age = today.year-recorded_date.year-((today.month, today.day) < (recorded_date.month, recorded_date.day))
        member_details.gender_id = member[3].value
        member_details.dob_bs = member[4].value
        member_details.mobile_num = member[5].value
        member_details.citizenship_num = member[6].value
        member_details.marital_status_id = member[7].value
        member_details.education_status_id = member[11].value
        member_details.education_level_id = member[12].value
        member_details.has_technical_training = member[14].value
        member_details.technical_skill_id = member[15].value if str(member[15].value).strip() else None
        member_details.main_occupation_id = member[19].value   # if member[19].value else None
        member_details.has_disability = member[25].value
        # member_details.is_married = member[25].value
        member_details.has_chronic_disease = member[28].value if str(member[28].value).strip() else None
        member_details.foreign_stay = member[38].value if str(member[38].value).strip() else None
        member_details.is_child_marriage = member[10].value if str(member[10].value).strip() else None
        member_details.is_vaccinated = member[35].value if str(member[35].value).strip() else None
        member_details.disability_type_id = member[26].value if str(member[26].value).strip() else None
        member_details.disability_card_id = member[27].value if str(member[27].value).strip() else None
        member_details.foreign_reason_id = member[39].value if str(member[39].value).strip() else None
        member_details.country_visited_id = member[41].value if str(member[41].value).strip() else None
        member_details.disease_name = member[29].value if str(member[29].value).strip() else None
        # member_details.vaccine_name_id = member[36].value
        # vaccine_name = member[38].value
        # vaccine = vaccine_name.split(',')
        # for v_id in vaccine:
        #     member_details = Member()
        #     member_details.hh = hh
        #     member_details.id_string = member[52].value
        #     member_details.vaccine_name_id = int(v_id)
        member_details.save()
    #     print(member[2].value)  # data-Household-info-Expire-Member-Details-Expire-Member-Gender
    #     print(member[3].value)  # data-Household-info-Expire-Member-Details-Expire-Member-Age
    #     print("House hold key = ", member[7].value)  # PARENT_KEY
    #     print("===================")
    #
        #  Save member here

    for i, house_type in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        house_detail = HouseDetail()
        id_string = house_type[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        house_detail.hh = hh
        house_detail.id_string = id_string
        house_detail.roof_type_id = house_type[22].value
        house_detail.house_num = house_type[11].value
        house_detail.build_year_bs = house_type[25].value
        house_detail.house_type_id = house_type[21].value
        # house_detail.map_pass = house_type[23].value
        house_detail.map_pass_date_bs = house_type[24].value if str(house_type[24].value).strip() else None
        house_detail.image = house_type[27].value
        house_detail.save()

    for i, animal in enumerate(animal_details.iter_rows(min_row=2)):  # for house sheet
        household_animal = HouseholdAnimal()
        hh_id_string = animal[2].value
        hh = Household.objects.filter(id_string=hh_id_string).first()
        household_animal.household = hh
        household_animal.id_string = animal[3].value
        household_animal.animal_type_id = animal[0].value
        household_animal.count = animal[1].value
        household_animal.save()

    for i, rent_member in enumerate(rent_details.iter_rows(min_row=2)):  # for house sheet
        rent = RentMember()

        rent.id_string = rent_member[6].value
        hh_id_string = rent_member[5].value
        hoh = Household.objects.filter(id_string=hh_id_string).first()
        rent.hh = hoh
        rent.rent_member_head = rent_member[0].value
        rent.no_of_member = rent_member[1].value
        rent.room_no = rent_member[3].value
        # rent.migrated_from = rent_member[2].value
        rent.reason_for_migration = rent_member[4].value
        rent.save()

    for i, business in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        business_type = HouseholdBusiness()
        id_string = business[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        business_type.household = hh
        business_type.id_string = id_string
        business_type.business_type_id = business[36].value
        business_type.business_place = business[39].value
        # business_type.is_registered = business[37].value
        business_type.money_invested = business[38].value
        business_type.save()

    for i, helper in enumerate(helper_details.iter_rows(min_row=2)):  # for house sheet
        household_helper = HouseholdHelper()
        id_string = helper[6].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_helper.household = hh
        household_helper.id_string = helper[7].value
        household_helper.child_name = helper[1].value
        household_helper.age = helper[2].value
        household_helper.gender_id = helper[3].value
        household_helper.start_date_bs = helper[5].value
        # household_helper.child_labor_status = helper[4].value
        household_helper.save()

    for i, land in enumerate(land_details.iter_rows(min_row=2)):  # for house sheet
        household_land = HouseholdLand()
        id_string = land[10].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_land.household = hh
        household_land.id_string = land[11].value
        household_land.land_type_id = land[2].value
        household_land.area1 = land[3].value
        household_land.area2 = land[4].value
        household_land.area3 = land[5].value
        household_land.area4 = land[6].value
        household_land.kitta_no = land[8].value
        household_land.save()

    for i, expenses in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        household_expenses = HouseholdExpenses()
        id_string = expenses[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_expenses.household = hh
        household_expenses.id_string = id_string
        household_expenses.education_expenses = expenses[81].value
        household_expenses.health_expenses = expenses[80].value
        household_expenses.food_expenses = expenses[79].value
        household_expenses.other_expenses = expenses[82].value
        household_expenses.save()

    for i, incomes in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        id_string = incomes[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        income_source_ = incomes[40].value
        income_source = income_source_.split(',')
        for i_id in income_source:
            household_income = HouseholdIncome()
            household_income.household = hh
            household_income.id_string = id_string
            household_income.source_id = int(i_id)
            household_income.amount = incomes[42].value
            household_income.start_date_bs = incomes[41].value
            household_income.save()

    for i, facility in enumerate(households.iter_rows(min_row=2)):  # for house sheet

        id_string = facility[0].value
        hh = Household.objects.filter(id_string=id_string).first()

        facility_type = facility[62].value
        facility_name = facility_type.split(',')
        for f_id in facility_name:
            household_facility = HouseholdFacility()
            household_facility.household = hh
            household_facility.id_string = id_string
            household_facility.name_id = int(f_id)
            household_facility.count = facility[63].value
            household_facility.save()

    for i, festival in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        id_string = festival[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        festival_type = festival[34].value
        festival_name = festival_type.split(',')
        for f_id in festival_name:
            household_festival = HouseholdFestival()
            household_festival.household = hh
            household_festival.id_string = id_string
            household_festival.festival_id = int(f_id)
            household_festival.save()

    for i, water in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        id_string = water[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_water = HouseholdWaterSource()
        household_water.household = hh
        household_water.id_string = id_string
        household_water.water_source_id = water[59].value
        household_water.distance = water[61].value
        household_water.time = water[60].value
        household_water.save()

    for i, fuel in enumerate(households.iter_rows(min_row=2)):  # for house sheet
        id_string = fuel[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_fuel = HouseholdFuel()
        household_fuel.household = hh
        household_fuel.id_string = id_string
        household_fuel.light_fuel_id = fuel[55].value
        household_fuel.cooking_fuel_id = fuel[57].value
        household_fuel.garbage_management_id = fuel[52].value
        household_fuel.save()

    for i, healthpost in enumerate(households.iter_rows(min_row=2)):
        household_healthpost = HouseholdHealthPost()
        id_string = healthpost[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_healthpost.household = hh
        household_healthpost.id_string = id_string
        household_healthpost.distance = healthpost[70].value
        household_healthpost.time = healthpost[71].value
        household_healthpost.save()

    for i, education in enumerate(households.iter_rows(min_row=2)):
        household_education = HouseholdEducation()
        household_education.id_string = education[0].value
        id_string = education[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_education.household = hh
        household_education.id_string = id_string
        household_education.primary_distance = education[64].value
        household_education.secondary_distance = education[65].value
        household_education.higher_secondary_distance = education[66].value
        household_education.save()

    for i, sewage in enumerate(households.iter_rows(min_row=2)):
        household_sewage = HouseholdSewage()
        id_string = sewage[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_sewage.household = hh
        household_sewage.id_string = id_string
        household_sewage.toilet_type_id = sewage[54].value
        household_sewage.sewage_type_id = sewage[53].value
        household_sewage.save()

    for i, transport in enumerate(households.iter_rows(min_row=2)):
        household_transport = HouseholdPublicTransport()
        id_string = transport[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_transport.household = hh
        household_transport.id_string = id_string
        household_transport.road_width = transport[67].value
        household_transport.distance = transport[68].value
        household_transport.possible_distance = transport[69].value  # time
        household_transport.save()

    for i, disaster in enumerate(households.iter_rows(min_row=2)):
        household_disaster = HouseholdNaturalDisaster()
        id_string = disaster[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_disaster.household = hh
        household_disaster.id_string = id_string
        household_disaster.disaster_id = disaster[74].value if str(disaster[74].value).strip() else None
        household_disaster.house_damage_status_id = disaster[75].value if str(disaster[75].value).strip() else None
        household_disaster.save()

    for i, infrastructure in enumerate(households.iter_rows(min_row=2)):
        household_infrastructure = HouseholdInfrastructure()
        id_string = infrastructure[0].value
        hh = Household.objects.filter(id_string=id_string).first()
        household_infrastructure.household = hh
        household_infrastructure.id_string = id_string
        household_infrastructure.light_fuel_id = infrastructure[55].value
        household_infrastructure.cooking_fuel_id = infrastructure[57].value
        household_infrastructure.toilet_type_id = infrastructure[54].value
        household_infrastructure.sewage_type_id = infrastructure[53].value
        household_infrastructure.water_source_id = infrastructure[59].value
        household_infrastructure.garbage_management_id = infrastructure[52].value
        household_infrastructure.disaster_id = infrastructure[74].value if str(infrastructure[74].value).strip() else None
        household_infrastructure.house_damage_status_id = infrastructure[75].value if str(infrastructure[75].value).strip() else None
        household_infrastructure.save()

    return redirect('import')

